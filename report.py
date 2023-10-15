from cProfile import label
from tkinter import Y, Image
from turtle import color, width
from openpyxl import load_workbook
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import plotly.express as px
from pyparsing import alphas
import matplotlib
import matplotlib.font_manager as fm
import matplotlib.ticker as ticker

from math import pi
from matplotlib.path import Path
from matplotlib.spines import Spine
from matplotlib.transforms import Affine2D

import openpyxl as op
from openpyxl.drawing.image import Image
from PIL import Image as piImage
from openpyxl.styles import Alignment

# 글꼴변경
plt.rc('font', family='NanumGothic')  # For Windows
matplotlib.rcParams['font.family'] = 'NanumGothic'


def fileUpload(fileName):
    fileData = pd.read_excel(f'{fileName}.xlsx', dtype='object')
    return fileData


########################## 도축함수##############################################################################도축함수##############################################################################
def abattStatistics(data, abattorYear, cowSex):
    sex = (data['성별'] == f'{cowSex}')
    year = (data['도축일자'].dt.strftime('%Y') == f'{abattorYear}')
    gGrade1pp = (data['육질등급'] == '1++') | (data['육질등급'] == '1+')
    wGradeA = (data['육량등급'] == 'A')
    wGradeB = (data['육량등급'] == 'B')
    wGradeC = (data['육량등급'] == 'C')

    sexAllCounters = data.loc[sex, '개체번호'].count()
    yearCounters = data.loc[sex & year, '개체번호'].count()

    if yearCounters != 0:
        CWT = round(data.loc[sex & year, '도체중'].mean(), 1)
        EMA = round(data.loc[sex & year, '등심단면적'].mean(), 1)
        BFT = round(data.loc[sex & year, '등지방두께'].mean(), 1)
        MAR = round(data.loc[sex & year, '근내지방도'].mean(), 1)
        abattorMonth = round(data.loc[sex & year, '도축개월령'].mean(), 1)

        Counters1pp = data.loc[sex & year & gGrade1pp, '육질등급'].count()
        CountersA = data.loc[sex & year & wGradeA, '육량등급'].count()
        CountersB = data.loc[sex & year & wGradeB, '육량등급'].count()
        CountersC = data.loc[sex & year & wGradeC, '육량등급'].count()

        percent1pp = round((Counters1pp / yearCounters) * 100, 1)
        percentA = round((CountersA / yearCounters) * 100, 1)
        percentB = round((CountersB / yearCounters) * 100, 1)
        percentC = round((CountersC / yearCounters) * 100, 1)
    else:
        CWT = 0
        EMA = 0
        BFT = 0
        MAR = 0
        abattorMonth = 0
        percent1pp = 0
        percentA = 0
        percentB = 0
        percentC = 0

    return sexAllCounters, yearCounters, CWT, EMA, BFT, MAR, percent1pp, percentA, percentB, percentC, abattorMonth


def sexAllCounters(data):
    sexM = data.loc[data['성별'] == '수', '개체번호'].count()
    sexF = data.loc[data['성별'] == '암', '개체번호'].count()
    sexSum = data['성별'].count()
    return sexM, sexF, sexSum


def abattMonthAllcounters(data, sex):
    AbattMonthAverage = data.loc[data['성별'] == f'{sex}', '도축개월령'].mean()
    AbattMonthMax = data.loc[data['성별'] == f'{sex}', '도축개월령'].max()
    AbattMonthMin = data.loc[data['성별'] == f'{sex}', '도축개월령'].min()
    return AbattMonthAverage,  AbattMonthMin, AbattMonthMax
########################## 도축함수##############################################################################도축함수##############################################################################

########################## 사육함수##############################################################################사육함수##############################################################################


def sbv(x, mean, std):
    result = (x-mean)/std
    return result


def selectionIndex(x, weight):
    result = x * weight
    return result


def pRank(x, max):
    result = x/max * 100
    return result


def grade(x):
    if x < 25:
        result = "A"
    elif 25 <= x < 50:
        result = "B"
    elif 50 <= x < 75:
        result = "C"
    elif 75 <= x <= 100:
        result = "D"
    else:
        result = float("nan")
    return result
########################## 사육함수##############################################################################사육함수##############################################################################


########################## 파일업로드##############################################################################파일업로드##############################################################################

print("사육 파일을 업로드합니다.")
breedList = fileUpload('사육')
breedList = breedList.astype({
    '개체번호': 'string',
    'KPN': 'string',
    '어미개체': 'string',
    '출생일자': 'string',
    '성별': 'string',
    '농가명': 'string'
})
print("사육 파일 업로드 완료")

print("KPN 파일을 업로드합니다.")
kpnList = fileUpload('KPN')
kpnList = kpnList.astype({
    'KPN': 'string',
    'CWT_EBV': 'float',
    'EMA_EBV': 'float',
    'BFT_EBV': 'float',
    'MAR_EBV': 'float'
})
print("KPN 파일 업로드 완료")

print("레퍼런스 파일을 업로드합니다.")
referenceList = fileUpload('레퍼런스')
referenceList = referenceList.astype({
    '개체번호': 'string',
    '01CWT_EBV': 'float',
    '02EMA_EBV': 'float',
    '03BFT_EBV': 'float',
    '04MAR_EBV': 'float'
}).replace(0, np.NaN)
print("레퍼런스 파일 업로드 완료")

print("도축파일을 업로드합니다.")
abattCowList = fileUpload('도축')
abattCowList = abattCowList.astype({
    '개체번호': 'string',
    '성별': 'string',
    '출생일자': 'datetime64[ns]',
    '도축일자': 'datetime64[ns]',
    '도체중': 'float',
    '등심단면적': 'float',
    '등지방두께': 'float',
    '근내지방도': 'float',
    '도축개월령': 'float',
    '육질등급': 'string',
    '육량등급': 'string',
    '육량지수': 'float',
    '농가명': 'string',
    '최초농가': 'string',
    '최종농가': 'string',
    '구분': 'string'
})
print("도축파일 업로드 완료")

print("도축전국통계 파일을 업로드합니다.")
abattAll = fileUpload('도축전체평균_거세')
abattAll = abattAll.astype({
    'CWT': 'float',
    'EMA': 'float',
    'BFT': 'float',
    'MAR': 'float'
})
abattAll = abattAll.round(2)
print("도축전국통계 파일 업로드 완료")


farmName = list(input("농가명을 입력하세요. (입력 예시 : 김xx,이xx) : ").split(","))
########################## 파일업로드##############################################################################파일업로드##############################################################################

################# 대상농가 데이터 분류##################################대상농가 데이터 분류##################################대상농가 데이터 분류#################

str_expr = "(최종농가 in @farmName)"
str_expr2 = "((최초농가 in  @farmName) and (최종농가 not in @farmName))"
sortFarm = "(농가명 in @farmName)"
breedInfo = breedList.query(sortFarm)
inSideInfo = abattCowList.query(str_expr)
outSideInfo = abattCowList.query(str_expr2)

################# 대상농가 데이터 분류##################################대상농가 데이터 분류##################################대상농가 데이터 분류#################

################# 도축##################################도축##################################도축##################################도축##################################도축#############################

inYearAbattStatistics = []
outYearAbattStatistics = []

bullInSideInfo1 = list(abattStatistics(inSideInfo, 2022, '수'))
bullInSideInfo2 = list(abattStatistics(inSideInfo, 2021, '수'))
bullInSideInfo3 = list(abattStatistics(inSideInfo, 2020, '수'))
bullInSideInfo4 = list(abattStatistics(inSideInfo, 2019, '수'))
bullInSideInfo5 = list(abattStatistics(inSideInfo, 2018, '수'))

inYearAbattStatistics.append(bullInSideInfo1)
inYearAbattStatistics.append(bullInSideInfo2)
inYearAbattStatistics.append(bullInSideInfo3)
inYearAbattStatistics.append(bullInSideInfo4)
inYearAbattStatistics.append(bullInSideInfo5)

cowInSideInfo1 = list(abattStatistics(inSideInfo, 2022, '암'))
cowInSideInfo2 = list(abattStatistics(inSideInfo, 2021, '암'))
cowInSideInfo3 = list(abattStatistics(inSideInfo, 2020, '암'))
cowInSideInfo4 = list(abattStatistics(inSideInfo, 2019, '암'))
cowInSideInfo5 = list(abattStatistics(inSideInfo, 2018, '암'))

inYearAbattStatistics.append(cowInSideInfo1)
inYearAbattStatistics.append(cowInSideInfo2)
inYearAbattStatistics.append(cowInSideInfo3)
inYearAbattStatistics.append(cowInSideInfo4)
inYearAbattStatistics.append(cowInSideInfo5)

bulloutSideInfo1 = list(abattStatistics(outSideInfo, 2022, '수'))
bulloutSideInfo2 = list(abattStatistics(outSideInfo, 2021, '수'))
bulloutSideInfo3 = list(abattStatistics(outSideInfo, 2020, '수'))
bulloutSideInfo4 = list(abattStatistics(outSideInfo, 2019, '수'))
bulloutSideInfo5 = list(abattStatistics(outSideInfo, 2018, '수'))

outYearAbattStatistics.append(bulloutSideInfo1)
outYearAbattStatistics.append(bulloutSideInfo2)
outYearAbattStatistics.append(bulloutSideInfo3)
outYearAbattStatistics.append(bulloutSideInfo4)
outYearAbattStatistics.append(bulloutSideInfo5)

cowoutSideInfo1 = list(abattStatistics(outSideInfo, 2022, '암'))
cowoutSideInfo2 = list(abattStatistics(outSideInfo, 2021, '암'))
cowoutSideInfo3 = list(abattStatistics(outSideInfo, 2020, '암'))
cowoutSideInfo4 = list(abattStatistics(outSideInfo, 2019, '암'))
cowoutSideInfo5 = list(abattStatistics(outSideInfo, 2018, '암'))

outYearAbattStatistics.append(cowoutSideInfo1)
outYearAbattStatistics.append(cowoutSideInfo2)
outYearAbattStatistics.append(cowoutSideInfo3)
outYearAbattStatistics.append(cowoutSideInfo4)
outYearAbattStatistics.append(cowoutSideInfo5)

AbattStatisticsHeader = ['출하두수총합', '연도별출하두수', 'CWT평균', 'EMA평균',
                         'BFT평균', 'MAR평균', '1+이상비율', 'A', 'B', 'C', '연도별도축개월령평균']
AbattStatisticsIndex = ['2022_수', '2021_수', '2020_수', '2019_수',
                        '2018_수', '2022_암', '2021_암', '2020_암', '2019_암', '2018_암']
inYearDf = pd.DataFrame(inYearAbattStatistics,
                        columns=AbattStatisticsHeader, index=AbattStatisticsIndex)
outYearDf = pd.DataFrame(outYearAbattStatistics,
                         columns=AbattStatisticsHeader, index=AbattStatisticsIndex)

################# 도축##################################도축##################################도축##################################도축##################################도축#############################

################################ 도축 컨텐츠 제작#############################################################도축 컨텐츠 제작#############################################################도축 컨텐츠 제작#############################

# page 1-1 농가현황

farmSituation = []

inAbattCount = list(sexAllCounters(inSideInfo))
outAbattCount = list(sexAllCounters(outSideInfo))
breedAbattCount = list(sexAllCounters(breedInfo))

farmSituation.append(outAbattCount)
farmSituation.append(inAbattCount)
farmSituation.append(breedAbattCount)

farmSituationDf = pd.DataFrame(
    farmSituation, index=['도축_외부', '도축_내부', '사육두수'], columns=['수', '암', '합계'])
farmSituationDf = farmSituationDf.transpose()  # 농가현황

# page 1-2 도축추세현황

abattSituation = []

bull = abattMonthAllcounters(inSideInfo, '수')
cow = abattMonthAllcounters(inSideInfo, '암')

abattSituation.append(bull)
abattSituation.append(cow)

abattSituationDf = pd.DataFrame(
    abattSituation, index=['수', '암'], columns=['평균개월', '최소개월', '최대개월'])
abattSituationDf = abattSituationDf.transpose().round(0)  # 도축추세현황
abattSituationDf.replace(np.NaN, 0, inplace=True)

# page 2-1 도축개월령현황

abattMonthAv = pd.DataFrame()
abattMonthAv['내부도축개월령'] = inYearDf['연도별도축개월령평균'].iloc[:5]
abattMonthAv['외부도축개월령'] = outYearDf['연도별도축개월령평균'].iloc[:5]
abattMonthAv = abattMonthAv.transpose()
abattMonthAv.replace(0, '-', inplace=True)
abattMonthAv = abattMonthAv[['2018_수', '2019_수',
                             '2020_수', '2021_수', '2022_수']]  # 도축개월령 현황


# page 2-1 도축개월령현황(그래프작성)
def bullAbattMonthScatter(indata, outdata, sex):
    inAbattMonth = indata[(indata['성별'] == '수') & (indata['도축개월령'])]
    inAbattMonth = inAbattMonth[['성별', '도축개월령']]
    inAbattMonth.replace(0, np.NaN, inplace=True)
    inAbattMonth.loc[(inAbattMonth.성별 == '수'), '성별'] = '2'

    outAbattMonth = outdata[(outdata['성별'] == '수') & (outdata['도축개월령'])]
    outAbattMonth = outAbattMonth[['성별', '도축개월령']]
    outAbattMonth.replace(0, np.NaN, inplace=True)
    outAbattMonth.loc[(outAbattMonth.성별 == '수'), '성별'] = '1'

    allAbattMonth = pd.concat([inAbattMonth, outAbattMonth], axis=0)
    allAbattMonth = allAbattMonth.astype({
        '성별': 'float',
        '도축개월령': 'float',
    })
    plt.figure(figsize=(3, 7))
    plt.scatter(allAbattMonth['성별'], allAbattMonth['도축개월령'], s=5)
    ax = plt.subplot()
    plt.xlim([0, 3])
    ax.xaxis.set_major_locator(ticker.MultipleLocator(1))
    # ax.yaxis.set_major_locator(ticker.MultipleLocator(5))
    inout = ['', '1', '2', '']
    index = np.arange(len(inout))
    plt.xticks(index, inout)
    plt.grid(True, alpha=0.2)
    plt.savefig(f'./chart/{sex}도축개월분포_chart.jpg', bbox_inches='tight')


def cowAbattMonthScatter(data, sex):
    abattMonth = data[(data['성별'] == '암') & (data['도축개월령'])]
    abattMonth = abattMonth[['성별', '도축개월령']]
    abattMonth.replace(0, np.NaN, inplace=True)

    abattMonth.loc[(abattMonth.성별 == '암'), '성별'] = '1'
    abattMonth = abattMonth.astype({
        '성별': 'float',
        '도축개월령': 'float',
    })
    plt.figure(figsize=(3, 7))
    plt.scatter(abattMonth['성별'], abattMonth['도축개월령'], s=5)
    ax = plt.subplot()
    plt.xlim([0, 2])
    ax.xaxis.set_major_locator(ticker.MultipleLocator(1))
    # ax.yaxis.set_major_locator(ticker.MultipleLocator(10))
    inout = ['', '2', '']
    index = np.arange(len(inout))
    plt.xticks(index, inout)
    plt.grid(True, alpha=0.2)
    plt.savefig(f'./chart/{sex}도축개월분포_chart.jpg', bbox_inches='tight')


inAbattMonthScatter = cowAbattMonthScatter(inSideInfo, '암')
allAbattMonthScatter = bullAbattMonthScatter(inSideInfo, outSideInfo, '수')

# page 2-2 도축개월령추세(그래프작성)

# 도축개월령추세


def abattLineChart(phenotype, ChartInterval, min, max):
    year = [2022, 2021, 2020, 2019, 2018]
    inAv = inYearDf[f'{phenotype}평균'].iloc[:5]
    inAv.replace(0, np.NaN, inplace=True)
    outAv = outYearDf[f'{phenotype}평균'].iloc[:5]
    outAv.replace(0, np.NaN, inplace=True)
    allAv = abattAll.loc[:, f'{phenotype}']

    inOutSum = pd.concat([inAv, outAv], axis=0)
    inOutAllSum = pd.concat([inOutSum, allAv], axis=0)
    inOutAllSumValue = inOutAllSum.values
    inOutAllMaxValue = round(np.nanmax(inOutAllSumValue), 0)
    inOutAllMinValue = round(np.nanmin(inOutAllSumValue), 0)

    plt.figure(figsize=(7, 4))
    plt.title(f'{phenotype}', fontsize=15)

    plt.plot(year, inAv, marker='o', label='도축추세',)
    plt.plot(year, outAv, marker='o', label='외부도축추세')
    plt.plot(year, allAv, '--', c='gray', alpha=0.5, label='전국도축추세')
    ax = plt.subplot()
    plt.ylim([inOutAllMinValue - min, inOutAllMaxValue + max])
    ax.yaxis.set_major_locator(ticker.MultipleLocator(ChartInterval))
    ax.set(xticks=[2022, 2021, 2020, 2019, 2018])
    plt.grid(True, axis='x', linestyle='dotted')
    plt.legend(loc='upper left')
    plt.savefig(f'./chart/{phenotype}_chart.jpg', bbox_inches='tight')


cwt = abattLineChart('CWT', 50, 50, 50)
ema = abattLineChart('EMA', 10, 20, 20)
bft = abattLineChart('BFT', 2, 4, 4)
mar = abattLineChart('MAR', 1, 2, 2)


# '1+이상비율','A','B','C'
def abattWgradeChart(data, type):
    abattGrade = pd.DataFrame()
    abattGrade['A비율'] = data['A'].iloc[:5]
    abattGrade['B비율'] = data['B'].iloc[:5]
    abattGrade['C비율'] = data['C'].iloc[:5]
    abattGrade.replace(0, np.NaN, inplace=True)
    abattGrade = abattGrade.sort_index(ascending=True)
    ax = abattGrade.plot(kind="bar", stacked=True, figsize=(
        7, 4), color=['#ED7D31', '#bfbfbf', '#FFC000'])
    for p in ax.patches:
        left, bottom, width, height = p.get_bbox().bounds
        if left == 0 or width == 0 or height == 0:
            pass
        else:
            ax.annotate(f'{round(height,1)}%', xy=(
                left+width/2, bottom+height/2), ha='center', va='center', color='white')
        # plt.sca(ax)
    year = ['2018', '2019', '2020', '2021', '2022']
    index = np.arange(len(year))
    plt.legend(loc='upper center', ncol=3, bbox_to_anchor=(0.5, 1))
    plt.xticks(index, year, rotation=0)
    plt.ylim([0, 100])
    plt.savefig(f'./chart/{type}_chart.jpg', bbox_inches='tight')


inWgrade = abattWgradeChart(inYearDf, '내부')
outWgrade = abattWgradeChart(outYearDf, '외부')

# page 3 출하 형질별 현황 표

# ['출하두수총합','연도별출하두수','CWT평균','EMA평균','BFT평균','MAR평균','1+이상비율','A','B','C']


def abattPhenoSituation(data, rawData):
    abattSituation = data[['연도별출하두수', 'CWT평균', 'EMA평균',
                           'BFT평균', 'MAR평균', '1+이상비율', 'A', 'B', 'C']].copy()
    abattSituation['1+이상비율'] = abattSituation['1+이상비율'].astype(str) + '%'
    abattSituation['A'] = abattSituation['A'].astype(str) + '%'
    abattSituation['B'] = abattSituation['B'].astype(str) + '%'
    abattSituation['C'] = abattSituation['C'].astype(str) + '%'
    abattSituation.replace(np.NaN, 0, inplace=True)
    abattSituation.replace(0, '0.0', inplace=True)

    sexList = ['수', '암']
    abattorYear = ['2018', '2019', '2020', '2021', '2022']
    addTotalList = []

    for i in sexList:

        sex = (rawData['성별'] == f'{i}')
        year = (rawData['도축일자'].dt.strftime('%Y') == f'{abattorYear[0]}') | (rawData['도축일자'].dt.strftime('%Y') == f'{abattorYear[1]}') | (rawData['도축일자'].dt.strftime(
            '%Y') == f'{abattorYear[2]}') | (rawData['도축일자'].dt.strftime('%Y') == f'{abattorYear[3]}') | (rawData['도축일자'].dt.strftime('%Y') == f'{abattorYear[4]}')
        gGrade1pp = (rawData['육질등급'] == '1++') | (rawData['육질등급'] == '1+')
        wGradeA = (rawData['육량등급'] == 'A')
        wGradeB = (rawData['육량등급'] == 'B')
        wGradeC = (rawData['육량등급'] == 'C')
        sexAllCounters = rawData.loc[sex & year, '개체번호'].count()

        addData = []

        if sexAllCounters != 0:
            CWT = round(rawData.loc[sex & year, '도체중'].mean(), 1)
            EMA = round(rawData.loc[sex & year, '등심단면적'].mean(), 1)
            BFT = round(rawData.loc[sex & year, '등지방두께'].mean(), 1)
            MAR = round(rawData.loc[sex & year, '근내지방도'].mean(), 1)

            Counters1pp = rawData.loc[sex & year & gGrade1pp, '육질등급'].count()
            CountersA = rawData.loc[sex & year & wGradeA, '육량등급'].count()
            CountersB = rawData.loc[sex & year & wGradeB, '육량등급'].count()
            CountersC = rawData.loc[sex & year & wGradeC, '육량등급'].count()

            percent1pp = round((Counters1pp / sexAllCounters) * 100, 1)
            percentA = round((CountersA / sexAllCounters) * 100, 1)
            percentB = round((CountersB / sexAllCounters) * 100, 1)
            percentC = round((CountersC / sexAllCounters) * 100, 1)
        else:
            CWT = 0
            EMA = 0
            BFT = 0
            MAR = 0
            percent1pp = 0
            percentA = 0
            percentB = 0
            percentC = 0

        addData.append(sexAllCounters)
        addData.append(CWT)
        addData.append(EMA)
        addData.append(BFT)
        addData.append(MAR)
        addData.append(percent1pp)
        addData.append(percentA)
        addData.append(percentB)
        addData.append(percentC)
        addTotalList.append(addData)

    # 총합
    sex = (rawData['성별'] == sexList[0]) | (rawData['성별'] == sexList[1])
    year = (rawData['도축일자'].dt.strftime('%Y') == f'{abattorYear[0]}') | (rawData['도축일자'].dt.strftime('%Y') == f'{abattorYear[1]}') | (rawData['도축일자'].dt.strftime(
        '%Y') == f'{abattorYear[2]}') | (rawData['도축일자'].dt.strftime('%Y') == f'{abattorYear[3]}') | (rawData['도축일자'].dt.strftime('%Y') == f'{abattorYear[4]}')
    gGrade1pp = (rawData['육질등급'] == '1++') | (rawData['육질등급'] == '1+')
    wGradeA = (rawData['육량등급'] == 'A')
    wGradeB = (rawData['육량등급'] == 'B')
    wGradeC = (rawData['육량등급'] == 'C')
    sexAllCounters = rawData.loc[sex & year, '개체번호'].count()

    addData = []

    if sexAllCounters != 0:
        CWT = round(rawData.loc[sex & year, '도체중'].mean(), 1)
        EMA = round(rawData.loc[sex & year, '등심단면적'].mean(), 1)
        BFT = round(rawData.loc[sex & year, '등지방두께'].mean(), 1)
        MAR = round(rawData.loc[sex & year, '근내지방도'].mean(), 1)

        Counters1pp = rawData.loc[sex & year & gGrade1pp, '육질등급'].count()
        CountersA = rawData.loc[sex & year & wGradeA, '육량등급'].count()
        CountersB = rawData.loc[sex & year & wGradeB, '육량등급'].count()
        CountersC = rawData.loc[sex & year & wGradeC, '육량등급'].count()

        percent1pp = round((Counters1pp / sexAllCounters) * 100, 1)
        percentA = round((CountersA / sexAllCounters) * 100, 1)
        percentB = round((CountersB / sexAllCounters) * 100, 1)
        percentC = round((CountersC / sexAllCounters) * 100, 1)
    else:
        CWT = 0
        EMA = 0
        BFT = 0
        MAR = 0
        percent1pp = 0
        percentA = 0
        percentB = 0
        percentC = 0

    addData.append(sexAllCounters)
    addData.append(CWT)
    addData.append(EMA)
    addData.append(BFT)
    addData.append(MAR)
    addData.append(percent1pp)
    addData.append(percentA)
    addData.append(percentB)
    addData.append(percentC)
    addTotalList.append(addData)

    addDf = pd.DataFrame(addTotalList, index=['수', '암', '합'], columns=[
                         '연도별출하두수', 'CWT평균', 'EMA평균', 'BFT평균', 'MAR평균', '1+이상비율', 'A', 'B', 'C'])

    addDf['1+이상비율'] = addDf['1+이상비율'].astype(str) + '%'
    addDf['A'] = addDf['A'].astype(str) + '%'
    addDf['B'] = addDf['B'].astype(str) + '%'
    addDf['C'] = addDf['C'].astype(str) + '%'
    addDf.replace(np.NaN, 0, inplace=True)
    addDf.replace(0, '0.0', inplace=True)

    abattSituationResult = pd.concat([abattSituation, addDf], axis=0)

    return abattSituationResult


inAbattPhenoSituation = abattPhenoSituation(inYearDf, inSideInfo)
outAbattPhenoSituation = abattPhenoSituation(outYearDf, outSideInfo)


################################ 도축 컨텐츠 제작#############################################################도축 컨텐츠 제작#############################################################도축 컨텐츠 제작#############################

################# 사육##################################사육##################################사육##################################사육##################################사육#############################

referenceStd = referenceList.std(numeric_only=True)
referenceMean = referenceList.mean(numeric_only=True)

# 전체 사육 계산
referenceList['01CWT_SBV'] = referenceList['01CWT_EBV'].map(
    lambda x: sbv(x, referenceMean['01CWT_EBV'], referenceStd['01CWT_EBV']))
referenceList['02EMA_SBV'] = referenceList['02EMA_EBV'].map(
    lambda x: sbv(x, referenceMean['02EMA_EBV'], referenceStd['02EMA_EBV']))
referenceList['03BFT_SBV'] = referenceList['03BFT_EBV'].map(
    lambda x: sbv(x, referenceMean['03BFT_EBV'], referenceStd['03BFT_EBV'])) * -1
referenceList['04MAR_SBV'] = referenceList['04MAR_EBV'].map(
    lambda x: sbv(x, referenceMean['04MAR_EBV'], referenceStd['04MAR_EBV']))

referenceSbvMean = referenceList.mean(numeric_only=True)  # 표준화 평균(차트용)

referenceListChart = referenceList.copy(deep=True)  # 컨설팅농가 유전능력 평균 현황 용

referenceList['01CWT_SBV_Rank'] = referenceList['01CWT_SBV'].rank(
    method='dense', ascending=False)
referenceList['02EMA_SBV_Rank'] = referenceList['02EMA_SBV'].rank(
    method='dense', ascending=False)
referenceList['03BFT_SBV_Rank'] = referenceList['03BFT_SBV'].rank(
    method='dense', ascending=False)
referenceList['04MAR_SBV_Rank'] = referenceList['04MAR_SBV'].rank(
    method='dense', ascending=False)
CWT_SBV_Rank_Max = referenceList['01CWT_SBV_Rank'].max()
EMA_SBV_Rank_Max = referenceList['02EMA_SBV_Rank'].max()
BFT_SBV_Rank_Max = referenceList['03BFT_SBV_Rank'].max()
MAR_SBV_Rank_Max = referenceList['04MAR_SBV_Rank'].max()
referenceList['01CWT_SBV_PRank'] = referenceList['01CWT_SBV_Rank'].map(
    lambda x: pRank(x, CWT_SBV_Rank_Max))
referenceList['02EMA_SBV_PRank'] = referenceList['02EMA_SBV_Rank'].map(
    lambda x: pRank(x, EMA_SBV_Rank_Max))
referenceList['03BFT_SBV_PRank'] = referenceList['03BFT_SBV_Rank'].map(
    lambda x: pRank(x, BFT_SBV_Rank_Max))
referenceList['04MAR_SBV_PRank'] = referenceList['04MAR_SBV_Rank'].map(
    lambda x: pRank(x, MAR_SBV_Rank_Max))
referenceList['01CWT_SBV_grade'] = referenceList['01CWT_SBV_PRank'].map(
    lambda x: grade(x))
referenceList['02EMA_SBV_grade'] = referenceList['02EMA_SBV_PRank'].map(
    lambda x: grade(x))
referenceList['03BFT_SBV_grade'] = referenceList['03BFT_SBV_PRank'].map(
    lambda x: grade(x))
referenceList['04MAR_SBV_grade'] = referenceList['04MAR_SBV_PRank'].map(
    lambda x: grade(x))


referenceList['CWT_Si'] = referenceList['01CWT_SBV'].map(lambda x: selectionIndex(x, 5)) + referenceList['02EMA_SBV'].map(lambda x: selectionIndex(
    x, 1)) + referenceList['03BFT_SBV'].map(lambda x: selectionIndex(x, 1)) + referenceList['04MAR_SBV'].map(lambda x: selectionIndex(x, 2))
referenceList['MAR_Si'] = referenceList['01CWT_SBV'].map(lambda x: selectionIndex(x, 1)) + referenceList['02EMA_SBV'].map(lambda x: selectionIndex(
    x, 1)) + referenceList['03BFT_SBV'].map(lambda x: selectionIndex(x, 1)) + referenceList['04MAR_SBV'].map(lambda x: selectionIndex(x, 4))
referenceList['CWT_Si_Rank'] = referenceList['CWT_Si'].rank(
    method='dense', ascending=False)
referenceList['MAR_Si_Rank'] = referenceList['MAR_Si'].rank(
    method='dense', ascending=False)
CWT_Si_Rank_Max = referenceList['CWT_Si_Rank'].max()
MAR_Si_Rank_Max = referenceList['MAR_Si_Rank'].max()
referenceList['CWT_Si_PRank'] = referenceList['CWT_Si_Rank'].map(
    lambda x: pRank(x, CWT_Si_Rank_Max))
referenceList['MAR_Si_PRank'] = referenceList['MAR_Si_Rank'].map(
    lambda x: pRank(x, MAR_Si_Rank_Max))


# 농가 사육 계산
breedResultDf = pd.merge(breedInfo, referenceList, how='left', on='개체번호')

farmReferenceSbvMean = breedResultDf.mean(numeric_only=True)  # 표준화 평균(차트용)

breedResultDf['01CWT_SBV_FRank'] = breedResultDf['01CWT_SBV'].rank(
    method='dense', ascending=False)
breedResultDf['02EMA_SBV_FRank'] = breedResultDf['02EMA_SBV'].rank(
    method='dense', ascending=False)
breedResultDf['03BFT_SBV_FRank'] = breedResultDf['03BFT_SBV'].rank(
    method='dense', ascending=False)
breedResultDf['04MAR_SBV_FRank'] = breedResultDf['04MAR_SBV'].rank(
    method='dense', ascending=False)
breedResultDf['CWT_Si_FRank'] = breedResultDf['CWT_Si'].rank(
    method='dense', ascending=False)
breedResultDf['MAR_Si_FRank'] = breedResultDf['MAR_Si'].rank(
    method='dense', ascending=False)

growthGQuality = breedResultDf.iloc[:, [
    0, 3, 1, 2, 4, 6, 32, 7, 33, 8, 34, 9, 35, 22, 25]]

growthImprove = growthGQuality[((growthGQuality['01CWT_SBV_grade'] == 'C') | (growthGQuality['01CWT_SBV_grade'] == 'D')) & (
    (growthGQuality['04MAR_SBV_grade'] == 'C') | (growthGQuality['04MAR_SBV_grade'] == 'D'))]
growthEnhence = growthGQuality[((growthGQuality['01CWT_SBV_grade'] == 'C') | (growthGQuality['01CWT_SBV_grade'] == 'D')) & (
    (growthGQuality['04MAR_SBV_grade'] == 'A') | (growthGQuality['04MAR_SBV_grade'] == 'B'))]
gQualityImprove = growthGQuality[((growthGQuality['01CWT_SBV_grade'] == 'A') | (growthGQuality['01CWT_SBV_grade'] == 'B')) & (
    (growthGQuality['04MAR_SBV_grade'] == 'C') | (growthGQuality['04MAR_SBV_grade'] == 'D'))]
gQualityEnhence = growthGQuality[((growthGQuality['01CWT_SBV_grade'] == 'A') | (growthGQuality['01CWT_SBV_grade'] == 'B')) & (
    (growthGQuality['04MAR_SBV_grade'] == 'A') | (growthGQuality['04MAR_SBV_grade'] == 'B'))]


growthImprove = growthImprove.sort_values(by=['01CWT_SBV_FRank'])
growthEnhence = growthEnhence.sort_values(by=['01CWT_SBV_FRank'])
gQualityImprove = gQualityImprove.sort_values(by=['04MAR_SBV_FRank'])
gQualityEnhence = gQualityEnhence.sort_values(by=['04MAR_SBV_FRank'])

growthImprove = growthImprove.fillna(0)
growthEnhence = growthEnhence.fillna(0)
gQualityImprove = gQualityImprove.fillna(0)
gQualityEnhence = gQualityEnhence.fillna(0)

growthImprove = growthImprove.round(2)
growthEnhence = growthEnhence.round(2)
gQualityImprove = gQualityImprove.round(2)
gQualityEnhence = gQualityEnhence.round(2)


# selectionIndexCwt = breedResultDf.iloc[:,[0,3,1,2,4,6,7,8,9,36,30]]
# selectionIndexMar = breedResultDf.iloc[:,[0,3,1,2,4,6,7,8,9,37,31]]


# selectionIndexCwt = selectionIndexCwt.sort_values(by=['CWT_Si_FRank'])
# selectionIndexMar = selectionIndexMar.sort_values(by=['MAR_Si_FRank'])

# selectionIndexCwt = selectionIndexCwt.fillna(0)
# selectionIndexMar = selectionIndexMar.fillna(0)

# selectionIndexCwt = selectionIndexCwt.round(1)
# selectionIndexMar = selectionIndexMar.round(1)

# selectionIndexCwt['CWT_Si_PRank'] = selectionIndexCwt['CWT_Si_PRank'].astype(str) + '%'
# print("도체중우선 선발지수sheet 완성")
# selectionIndexMar['MAR_Si_PRank'] = selectionIndexMar['MAR_Si_PRank'].astype(str) + '%'
# print("근내지방도우선 선발지수sheet 완성")
##################################################################################################################################################################################################################
# 농가 사육 계산 (암수 구분)
# 암

FsexSelect = (breedInfo['성별'] == '암')
FbreedInfo = breedInfo.loc[FsexSelect].copy(deep=True)

FbreedResultDf = pd.merge(FbreedInfo, referenceList, how='left', on='개체번호')


FbreedResultDf['01CWT_SBV_FRank'] = FbreedResultDf['01CWT_SBV'].rank(
    method='dense', ascending=False)
FbreedResultDf['02EMA_SBV_FRank'] = FbreedResultDf['02EMA_SBV'].rank(
    method='dense', ascending=False)
FbreedResultDf['03BFT_SBV_FRank'] = FbreedResultDf['03BFT_SBV'].rank(
    method='dense', ascending=False)
FbreedResultDf['04MAR_SBV_FRank'] = FbreedResultDf['04MAR_SBV'].rank(
    method='dense', ascending=False)
FbreedResultDf['CWT_Si_FRank'] = FbreedResultDf['CWT_Si'].rank(
    method='dense', ascending=False)
FbreedResultDf['MAR_Si_FRank'] = FbreedResultDf['MAR_Si'].rank(
    method='dense', ascending=False)

FgrowthGQuality = FbreedResultDf.iloc[:, [
    0, 3, 1, 2, 4, 6, 32, 7, 33, 8, 34, 9, 35, 22, 25]]

FgrowthImprove = FgrowthGQuality[((FgrowthGQuality['01CWT_SBV_grade'] == 'C') | (FgrowthGQuality['01CWT_SBV_grade'] == 'D')) & (
    (FgrowthGQuality['04MAR_SBV_grade'] == 'C') | (FgrowthGQuality['04MAR_SBV_grade'] == 'D'))]
FgrowthEnhence = FgrowthGQuality[((FgrowthGQuality['01CWT_SBV_grade'] == 'C') | (FgrowthGQuality['01CWT_SBV_grade'] == 'D')) & (
    (FgrowthGQuality['04MAR_SBV_grade'] == 'A') | (FgrowthGQuality['04MAR_SBV_grade'] == 'B'))]
FgQualityImprove = FgrowthGQuality[((FgrowthGQuality['01CWT_SBV_grade'] == 'A') | (FgrowthGQuality['01CWT_SBV_grade'] == 'B')) & (
    (FgrowthGQuality['04MAR_SBV_grade'] == 'C') | (FgrowthGQuality['04MAR_SBV_grade'] == 'D'))]
FgQualityEnhence = FgrowthGQuality[((FgrowthGQuality['01CWT_SBV_grade'] == 'A') | (FgrowthGQuality['01CWT_SBV_grade'] == 'B')) & (
    (FgrowthGQuality['04MAR_SBV_grade'] == 'A') | (FgrowthGQuality['04MAR_SBV_grade'] == 'B'))]


FgrowthImprove = FgrowthImprove.sort_values(by=['01CWT_SBV_FRank'])
FgrowthEnhence = FgrowthEnhence.sort_values(by=['01CWT_SBV_FRank'])
FgQualityImprove = FgQualityImprove.sort_values(by=['04MAR_SBV_FRank'])
FgQualityEnhence = FgQualityEnhence.sort_values(by=['04MAR_SBV_FRank'])

FgrowthImprove = FgrowthImprove.fillna(0)
FgrowthEnhence = FgrowthEnhence.fillna(0)
FgQualityImprove = FgQualityImprove.fillna(0)
FgQualityEnhence = FgQualityEnhence.fillna(0)

FgrowthImprove = FgrowthImprove.round(2)
print("암_성장개량sheet완성")
FgrowthEnhence = FgrowthEnhence.round(2)
print("암_성장강화sheet완성")
FgQualityImprove = FgQualityImprove.round(2)
print("암_육질개량sheet완성")
FgQualityEnhence = FgQualityEnhence.round(2)
print("암_육질강화sheet완성")


FselectionIndexCwt = FbreedResultDf.iloc[:, [
    0, 3, 1, 2, 4, 6, 7, 8, 9, 36, 30]]
FselectionIndexMar = FbreedResultDf.iloc[:, [
    0, 3, 1, 2, 4, 6, 7, 8, 9, 37, 31]]


FselectionIndexCwt = FselectionIndexCwt.sort_values(by=['CWT_Si_FRank'])
FselectionIndexMar = FselectionIndexMar.sort_values(by=['MAR_Si_FRank'])

FselectionIndexCwt = FselectionIndexCwt.fillna(0)
FselectionIndexMar = FselectionIndexMar.fillna(0)

FselectionIndexCwt = FselectionIndexCwt.round(2)
FselectionIndexMar = FselectionIndexMar.round(2)

FselectionIndexCwt['CWT_Si_PRank'] = FselectionIndexCwt['CWT_Si_PRank'].astype(
    str) + '%'
print("암_도체중우선 선발지수sheet 완성")
FselectionIndexMar['MAR_Si_PRank'] = FselectionIndexMar['MAR_Si_PRank'].astype(
    str) + '%'
print("암_근내지방도우선 선발지수sheet 완성")

################

# 수

MsexSelect = (breedInfo['성별'] == '수')
MbreedInfo = breedInfo.loc[MsexSelect].copy(deep=True)

MbreedResultDf = pd.merge(MbreedInfo, referenceList, how='left', on='개체번호')


MbreedResultDf['01CWT_SBV_FRank'] = MbreedResultDf['01CWT_SBV'].rank(
    method='dense', ascending=False)
MbreedResultDf['02EMA_SBV_FRank'] = MbreedResultDf['02EMA_SBV'].rank(
    method='dense', ascending=False)
MbreedResultDf['03BFT_SBV_FRank'] = MbreedResultDf['03BFT_SBV'].rank(
    method='dense', ascending=False)
MbreedResultDf['04MAR_SBV_FRank'] = MbreedResultDf['04MAR_SBV'].rank(
    method='dense', ascending=False)
MbreedResultDf['CWT_Si_FRank'] = MbreedResultDf['CWT_Si'].rank(
    method='dense', ascending=False)
MbreedResultDf['MAR_Si_FRank'] = MbreedResultDf['MAR_Si'].rank(
    method='dense', ascending=False)

MgrowthGQuality = MbreedResultDf.iloc[:, [
    0, 3, 1, 2, 4, 6, 32, 7, 33, 8, 34, 9, 35, 22, 25]]

MgrowthImprove = MgrowthGQuality[((MgrowthGQuality['01CWT_SBV_grade'] == 'C') | (MgrowthGQuality['01CWT_SBV_grade'] == 'D')) & (
    (MgrowthGQuality['04MAR_SBV_grade'] == 'C') | (MgrowthGQuality['04MAR_SBV_grade'] == 'D'))]
MgrowthEnhence = MgrowthGQuality[((MgrowthGQuality['01CWT_SBV_grade'] == 'C') | (MgrowthGQuality['01CWT_SBV_grade'] == 'D')) & (
    (MgrowthGQuality['04MAR_SBV_grade'] == 'A') | (MgrowthGQuality['04MAR_SBV_grade'] == 'B'))]
MgQualityImprove = MgrowthGQuality[((MgrowthGQuality['01CWT_SBV_grade'] == 'A') | (MgrowthGQuality['01CWT_SBV_grade'] == 'B')) & (
    (MgrowthGQuality['04MAR_SBV_grade'] == 'C') | (MgrowthGQuality['04MAR_SBV_grade'] == 'D'))]
MgQualityEnhence = MgrowthGQuality[((MgrowthGQuality['01CWT_SBV_grade'] == 'A') | (MgrowthGQuality['01CWT_SBV_grade'] == 'B')) & (
    (MgrowthGQuality['04MAR_SBV_grade'] == 'A') | (MgrowthGQuality['04MAR_SBV_grade'] == 'B'))]


MgrowthImprove = MgrowthImprove.sort_values(by=['01CWT_SBV_FRank'])
MgrowthEnhence = MgrowthEnhence.sort_values(by=['01CWT_SBV_FRank'])
MgQualityImprove = MgQualityImprove.sort_values(by=['04MAR_SBV_FRank'])
MgQualityEnhence = MgQualityEnhence.sort_values(by=['04MAR_SBV_FRank'])

MgrowthImprove = MgrowthImprove.fillna(0)
MgrowthEnhence = MgrowthEnhence.fillna(0)
MgQualityImprove = MgQualityImprove.fillna(0)
MgQualityEnhence = MgQualityEnhence.fillna(0)

MgrowthImprove = MgrowthImprove.round(2)
print("수_성장개량sheet완성")
MgrowthEnhence = MgrowthEnhence.round(2)
print("수_성장강화sheet완성")
MgQualityImprove = MgQualityImprove.round(2)
print("수_육질개량sheet완성")
MgQualityEnhence = MgQualityEnhence.round(2)
print("수_육질강화sheet완성")


MselectionIndexCwt = MbreedResultDf.iloc[:, [
    0, 3, 1, 2, 4, 6, 7, 8, 9, 36, 30]]
MselectionIndexMar = MbreedResultDf.iloc[:, [
    0, 3, 1, 2, 4, 6, 7, 8, 9, 37, 31]]


MselectionIndexCwt = MselectionIndexCwt.sort_values(by=['CWT_Si_FRank'])
MselectionIndexMar = MselectionIndexMar.sort_values(by=['MAR_Si_FRank'])

MselectionIndexCwt = MselectionIndexCwt.fillna(0)
MselectionIndexMar = MselectionIndexMar.fillna(0)

MselectionIndexCwt = MselectionIndexCwt.round(2)
MselectionIndexMar = MselectionIndexMar.round(2)

MselectionIndexCwt['CWT_Si_PRank'] = MselectionIndexCwt['CWT_Si_PRank'].astype(
    str) + '%'
print("수_도체중우선 선발지수sheet 완성")
MselectionIndexMar['MAR_Si_PRank'] = MselectionIndexMar['MAR_Si_PRank'].astype(
    str) + '%'
print("수_근내지방도우선 선발지수sheet 완성")

##################################################################################################################################################################################################################
# 현황 카운트
# #농가현황

breedTypeTable = []
growthImproveCount = list(sexAllCounters(growthImprove))
growthEnhenceCount = list(sexAllCounters(growthEnhence))
gQualityImproveCount = list(sexAllCounters(gQualityImprove))
gQualityEnhenceCount = list(sexAllCounters(gQualityEnhence))
breedTypeTable.append(growthImproveCount)
breedTypeTable.append(growthEnhenceCount)
breedTypeTable.append(gQualityImproveCount)
breedTypeTable.append(gQualityEnhenceCount)

breedTypeDf = pd.DataFrame(breedTypeTable, columns=['암', '수', '합'])


breedTypeSum = breedTypeDf['합'].sum()

breedTypeDf['분포율'] = round(breedTypeDf['합']/breedTypeSum * 100, 1)

sumData = {
    '암': [breedTypeDf['암'].sum()],
    '수': [breedTypeDf['수'].sum()],
    '합': [breedTypeDf['합'].sum()],
    '분포율': [100]
}
sumDf = pd.DataFrame(sumData)

breedTyperesult = pd.concat([breedTypeDf, sumDf])

breedTyperesult.index = ['도체약 육질약', '도체약 육질강', '도체강 육질약', '도체강 육질강', '총합']

breedPiChartSeries = breedTyperesult.loc[[
    '도체약 육질약', '도체약 육질강', '도체강 육질약', '도체강 육질강'], ['분포율']]  # 파이 차트용 데이터

breedTyperesult['분포율'] = breedTyperesult['분포율'].astype(str) + '%'

breedTyperesult = breedTyperesult[['수', '암', '합', '분포율']]
breedTyperesult.columns = ['암', '수', '합', '분포율']

# breedTyperesult 개체별 유전능력 유형 분류
# breedTyperesult 분석결과 0인 애들은 제외되고 카운팅 된다.

## 파이차트##


def breedPiChart(dataSeries, type):
    breedPiChartTable = pd.DataFrame(dataSeries, columns=['분포율'])
    breedPiChartTable = breedPiChartTable.transpose()
    colors = ['#5B9BD5', '#ED7D31', '#A5A5A5', '#ffc000']
    wedgeprops = {'width': 0.6, 'edgecolor': 'w', 'linewidth': 5}
    labels = ['도체약 육질약', '도체약 육질강', '도체강 육질약', '도체강 육질강']
    plt.cla()
    plt.pie(breedPiChartTable.loc['분포율'], autopct='%.1f%%', startangle=260,
            counterclock=False, colors=colors, wedgeprops=wedgeprops)
    plt.legend(loc='lower center', labels=labels, ncol=4, fontsize=10)
    plt.savefig(f'./chart/{type}_chart.jpg', bbox_inches='tight')


breedPiChart(breedPiChartSeries, '개체별유형분류')


################ 사육chart##################################사육chart##################################사육chart##################################사육chart##################################사육chart#############################

# 레이더 차트

breedChartSeries = pd.concat([referenceSbvMean, farmReferenceSbvMean], axis=1)
breedSeedDf = pd.DataFrame(
    index=['01CWT_SBV', '02EMA_SBV', '03BFT_SBV', '04MAR_SBV'])
breedChartDf = pd.concat([breedSeedDf, breedChartSeries], axis=1)
breedChartDf.columns = ['전체평균', '농가평균']
breedChartTable = breedChartDf.loc[[
    '01CWT_SBV', '02EMA_SBV', '03BFT_SBV', '04MAR_SBV'], ['전체평균', '농가평균']]  # 차트용(SBV)

breedTable = round(breedChartDf.loc[['01CWT_EBV', '02EMA_EBV', '03BFT_EBV', '04MAR_EBV'], [
                   '전체평균', '농가평균']], 2)  # 보고서용(EBV) # 유전능력 현황


breedChartTable = breedChartTable.transpose()
breedChartTable = breedChartTable.reset_index()
breedChartTable.columns = ['Character',
                           'CWT_SBV', 'EMA_SBV', 'BFT_SBV', 'MAR_SBV']

breedRaderMax = breedChartTable.max(numeric_only=True, axis=1)
breedRaderMin = breedChartTable.min(numeric_only=True, axis=1)
breedRaderSum = pd.concat([breedRaderMax, breedRaderMin], axis=0)
breedRaderSumValue = breedRaderSum.values
breedRaderMaxValue = np.max(breedRaderSumValue)
breedRaderMinValue = np.min(breedRaderSumValue)


def raderchart(df, min, max, type, mingrid, maxgrid):

    labels = df.columns[1:]
    num_labels = len(labels)

    angles = [x/float(num_labels)*(2*pi) for x in range(num_labels)]  # 각 등분점
    angles += angles[:1]  # 시작점으로 다시 돌아와야하므로 시작점 추가
    # my_palette = plt.cm.get_cmap("Set2", len(df.index))
    # mycolor = ['#FD4B04','#95A3A6'] #['전체','농가']
    mycolor = ['#00175A', '#ED7625']
    myalpha = [0.8, 0.15]
    mylinestyle = ['dashed', 'solid']
    fig = plt.figure(figsize=(5, 5))
    fig.set_facecolor('white')
    ax = fig.add_subplot(polar=True)
    for i, row in df.iterrows():
        color = mycolor[i]
        data = df.iloc[i].drop('Character').tolist()
        data += data[:1]

        ax.set_theta_offset(pi / 2)  # 시작점
        ax.set_theta_direction(-1)  # 그려지는 방향 시계방향

        phenoName = ['CWT', 'EMA', 'BFT', 'MAR']
        plt.xticks(angles[:-1], phenoName, fontsize=13)  # x축 눈금 라벨
        ax.tick_params(axis='x', which='major', pad=15)  # x축과 눈금 사이에 여백을 준다.
        ax.set_rlabel_position(0)  # y축 각도 설정(degree 단위)
        ax.axes.yaxis.set_ticklabels([])
        plt.ylim(min-mingrid, max+maxgrid)

        ax.plot(angles, data, color=color, linewidth=3,
                linestyle=mylinestyle[i], label=row.Character)  # 레이더 차트 출력
        # ax.fill(angles, data, color=color, alpha=myalpha[i]) ## 도형 안쪽에 색을 채워준다.

    plt.legend(loc=(0.9, 0.9))
    plt.savefig(f'./chart/{type}rader_chart.jpg', bbox_inches='tight')


breedRaderChart = raderchart(breedChartTable, breedRaderMinValue,
                             breedRaderMaxValue, '사육', mingrid=0.5, maxgrid=0.5)  # 사육 레이더 차트

# 백분율 차트

referenceListChart = referenceListChart[[
    '01CWT_SBV', '02EMA_SBV', '03BFT_SBV', '04MAR_SBV']]
referenceListChart['구분'] = '레퍼런스'

farmMeanData = breedChartDf.loc[[
    '01CWT_SBV', '02EMA_SBV', '03BFT_SBV', '04MAR_SBV'], ['농가평균']]  # 차트용(SBV)

farmMeanDf = pd.DataFrame(farmMeanData)
farmMeanDf = farmMeanDf.transpose()
farmMeanDf['구분'] = '농가평균'


farmMeanDfresult = pd.concat([referenceListChart, farmMeanDf])


farmMeanDfresult['01CWT_SBV_Rank'] = farmMeanDfresult['01CWT_SBV'].rank(
    method='dense', ascending=False)
farmMeanDfresult['02EMA_SBV_Rank'] = farmMeanDfresult['02EMA_SBV'].rank(
    method='dense', ascending=False)
farmMeanDfresult['03BFT_SBV_Rank'] = farmMeanDfresult['03BFT_SBV'].rank(
    method='dense', ascending=False)
farmMeanDfresult['04MAR_SBV_Rank'] = farmMeanDfresult['04MAR_SBV'].rank(
    method='dense', ascending=False)

Chart_CWT_SBV_Rank_Max = farmMeanDfresult['01CWT_SBV_Rank'].max()
Chart_EMA_SBV_Rank_Max = farmMeanDfresult['02EMA_SBV_Rank'].max()
Chart_BFT_SBV_Rank_Max = farmMeanDfresult['03BFT_SBV_Rank'].max()
Chart_MAR_SBV_Rank_Max = farmMeanDfresult['04MAR_SBV_Rank'].max()
farmMeanDfresult['01CWT_SBV_PRank'] = farmMeanDfresult['01CWT_SBV_Rank'].map(
    lambda x: pRank(x, Chart_CWT_SBV_Rank_Max))
farmMeanDfresult['02EMA_SBV_PRank'] = farmMeanDfresult['02EMA_SBV_Rank'].map(
    lambda x: pRank(x, Chart_EMA_SBV_Rank_Max))
farmMeanDfresult['03BFT_SBV_PRank'] = farmMeanDfresult['03BFT_SBV_Rank'].map(
    lambda x: pRank(x, Chart_BFT_SBV_Rank_Max))
farmMeanDfresult['04MAR_SBV_PRank'] = farmMeanDfresult['04MAR_SBV_Rank'].map(
    lambda x: pRank(x, Chart_MAR_SBV_Rank_Max))


farmMeanChartResult = farmMeanDfresult.loc[['농가평균'], [
    '01CWT_SBV_PRank', '02EMA_SBV_PRank', '03BFT_SBV_PRank', '04MAR_SBV_PRank']]
farmMeanChartResult = farmMeanChartResult.reset_index()


def percentRankChart(data, columnname, type):
    plt.cla()
    plt.figure(figsize=(3, 7))
    plt.scatter(1, data[f'{columnname}_SBV_PRank'], s=60, color='#ED7625')
    ax = plt.subplot()
    plt.xlim([0, 2])
    plt.ylim([0, 100])
    ax.invert_yaxis()
    ax.xaxis.set_major_locator(ticker.MultipleLocator(1))
    # ax.yaxis.set_major_locator(ticker.MultipleLocator(5))
    xinout = [f'{type}']
    xindex = [1]
    plt.xticks(xindex, xinout)

    ax.yaxis.set_major_locator(ticker.MultipleLocator(10))
    yinout = ['0%', '10%', '20%', '30%', '40%',
              '50%', '60%', '70%', '80%', '90%', '100%',]
    yindex = [0, 10, 20, 30, 40, 50, 60, 70, 80, 90, 100]
    plt.yticks(yindex, yinout)

    plt.grid(True, alpha=0.5)
    plt.savefig(f'./chart/{type}전체순위백분율분포_chart.jpg', bbox_inches='tight')


percentRankChartCwt = percentRankChart(farmMeanChartResult, '01CWT', 'CWT')
percentRankChartEma = percentRankChart(farmMeanChartResult, '02EMA', 'EMA')
percentRankChartBft = percentRankChart(farmMeanChartResult, '03BFT', 'BFT')
percentRankChartMar = percentRankChart(farmMeanChartResult, '04MAR', 'MAR')
################ 사육chart##################################사육chart##################################사육chart##################################사육chart##################################사육chart#############################


############# KPN############################KPN############################KPN############################KPN############################KPN############################KPN##############
KPN = breedInfo.value_counts(subset=['KPN']).to_frame()


kpnStd = kpnList.std(numeric_only=True)
kpnMean = kpnList.mean(numeric_only=True)
kpnList['CWT_SBV'] = kpnList['CWT_EBV'].map(
    lambda x: sbv(x, kpnMean['CWT_EBV'], kpnStd['CWT_EBV']))
kpnList['EMA_SBV'] = kpnList['EMA_EBV'].map(
    lambda x: sbv(x, kpnMean['EMA_EBV'], kpnStd['EMA_EBV']))
kpnList['BFT_SBV'] = kpnList['BFT_EBV'].map(
    lambda x: sbv(x, kpnMean['BFT_EBV'], kpnStd['BFT_EBV'])) * -1
kpnList['MAR_SBV'] = kpnList['MAR_EBV'].map(
    lambda x: sbv(x, kpnMean['MAR_EBV'], kpnStd['MAR_EBV']))
kpnList = kpnList.astype({
    'CWT_SBV': 'float',
    'EMA_SBV': 'float',
    'BFT_SBV': 'float',
    'MAR_SBV': 'float'
})
kpnList = kpnList.set_index('KPN')


kpndf = pd.DataFrame(KPN)
kpndf = kpndf.reset_index()
kpndf.rename(columns={0: 'count'}, inplace=True)
if (kpndf['KPN'] == 'unknown').any():
    kpndf = kpndf[kpndf.KPN != 'unknown']
else:
    pass

kpnResultDf = pd.merge(kpndf, kpnList, how='left', on='KPN')
kpnResultDf = kpnResultDf.astype({
    'count': 'float',
    'CWT_EBV': 'float',
    'EMA_EBV': 'float',
    'BFT_EBV': 'float',
    'MAR_EBV': 'float',
    'CWT_SBV': 'float',
    'EMA_SBV': 'float',
    'BFT_SBV': 'float',
    'MAR_SBV': 'float'
})
kpnSum = kpnResultDf['count'].sum()
kpnResultDf = kpnResultDf.round(2)
kpnResultDf['kpn%'] = round(
    kpnResultDf['count'].map(lambda x: pRank(x, kpnSum)), 1)
kpnResultDf['kpn%'] = kpnResultDf['kpn%'].astype(str) + '%'

kpnDistribution = kpnResultDf.iloc[:, [0, 1, 10, 2, 3, 4, 5]]

kpnDistribution = kpnDistribution.fillna(0)  # kpn 빈도 테이블

# KPN 빈도수 chart


def kpnDistributionChart(data):
    kpnCount = data[['KPN', 'count']]
    columnCount = kpnCount['KPN'].count()

    if columnCount > 23:
        kpnCount = kpnCount.loc[:22]
        plt.cla()
        plt.figure(figsize=(15, 7))
        plt.bar(kpnCount['KPN'], kpnCount['count'], color=['#EC7B2F'])

        for i, v in enumerate(kpnCount['KPN']):

            plt.text(v, kpnCount['count'][i], kpnCount['count'][i],
                     fontsize=9,
                     color='black',
                     horizontalalignment='center',
                     verticalalignment='bottom')

        plt.xticks(rotation=90)
        # plt.subplots_adjust(left = 0, bottom = 0, right = 1, top = 1, hspace = 0, wspace = 0)
        plt.savefig(f'./chart/KPN빈도_chart.jpg', bbox_inches='tight')

    else:
        plt.cla()
        plt.figure(figsize=(15, 7))
        plt.bar(kpnCount['KPN'], kpnCount['count'], color=['#EC7B2F'])
        for i, v in enumerate(kpnCount['KPN']):

            plt.text(v, kpnCount['count'][i], kpnCount['count'][i],
                     fontsize=9,
                     color='black',
                     horizontalalignment='center',
                     verticalalignment='bottom')

        plt.xticks(rotation=90)
        plt.savefig(f'./chart/KPN빈도_chart.jpg', bbox_inches='tight')


kpnDChart = kpnDistributionChart(kpnDistribution)  # kpn 사용 빈도

# kpn 유전능력 현황

allKpnSbvMean = kpnList.mean(numeric_only=True)
kpnResultDf2 = pd.merge(kpndf, kpnList, how='left', on='KPN')
kpnResultDf2 = kpnResultDf2.astype({
    'count': 'float',
    'CWT_EBV': 'float',
    'EMA_EBV': 'float',
    'BFT_EBV': 'float',
    'MAR_EBV': 'float',
    'CWT_SBV': 'float',
    'EMA_SBV': 'float',
    'BFT_SBV': 'float',
    'MAR_SBV': 'float'
})
kpnResultDf2['CWT_sum'] = kpnResultDf2['CWT_EBV'] * kpnResultDf2['count']
kpnResultDf2['EMA_sum'] = kpnResultDf2['EMA_EBV'] * kpnResultDf2['count']
kpnResultDf2['BFT_sum'] = kpnResultDf2['BFT_EBV'] * kpnResultDf2['count']
kpnResultDf2['MAR_sum'] = kpnResultDf2['MAR_EBV'] * kpnResultDf2['count']


farmKpnSbvMean = kpnResultDf2.mean(numeric_only=True)
kpnRaderDf = pd.concat([allKpnSbvMean, farmKpnSbvMean], axis=1)
kpnRaderDf = kpnRaderDf.loc[['CWT_SBV', 'EMA_SBV', 'BFT_SBV', 'MAR_SBV'], :]
kpnRaderDf = kpnRaderDf.transpose()
kpnRaderDf.index = ['전체평균', '농가평균']
kpnRaderDf = kpnRaderDf.reset_index()
kpnRaderDf.columns = ['Character', 'CWT_SBV', 'EMA_SBV', 'BFT_SBV', 'MAR_SBV']


# kpn 레이터 차트(함수 사육 레디어 차트 부분에 있음)

kpnRaderMax = kpnRaderDf.max(numeric_only=True, axis=1)
kpnRaderMin = kpnRaderDf.min(numeric_only=True, axis=1)
kpnRaderSum = pd.concat([kpnRaderMax, kpnRaderMin], axis=0)
kpnRaderSumValue = kpnRaderSum.values
kpnRaderMaxValue = np.max(kpnRaderSumValue)
kpnRaderMinValue = np.min(kpnRaderSumValue)


kpnRaderChart = raderchart(kpnRaderDf, kpnRaderMinValue,
                           kpnRaderMaxValue, 'kpn', mingrid=0.5, maxgrid=0.1)  # kpn 차트


kpnCountSum = kpnResultDf2['count'].sum()
kpnEbvSumDf = kpnResultDf2[['CWT_sum', 'EMA_sum', 'BFT_sum', 'MAR_sum']]
kpnEbvAllSum = kpnEbvSumDf.sum(axis=0)
kpnEbvMeanDf = pd.DataFrame(kpnEbvAllSum, columns=['Sum'])
kpnEbvMeanDf['Mean'] = round(kpnEbvMeanDf['Sum'] / kpnCountSum, 2)
kpnEbvMeanSeries = kpnEbvMeanDf['Mean']

kpnEbvMeanResult = pd.DataFrame(
    kpnEbvMeanSeries, columns=['Mean'])  # 씨수소 유전능력평균


############# KPN############################KPN############################KPN############################KPN############################KPN############################KPN##############


############################## 파일생성#######################################################################파일생성########################################

print("보고서 엑셀 생성 중 ........")


def fileDownload(df1, df2, df3, df4, df5, df6, df7, df8, df9, df10, df11, df12, df13, df14, df15, df16, df17, df18, df19, df20, df21):
    with pd.ExcelWriter(f'./reportResult/{farmName[0]}_농장보고서.xlsx') as writer:
        df1.to_excel(writer, sheet_name="농가&도축현황", startrow=1,
                     startcol=1)  # farmSituationDf 농가 현황
        df2.to_excel(writer, sheet_name="농가&도축현황", startrow=8,
                     startcol=1)  # abattSituationDf 도축 추세현황
        df3.to_excel(writer, sheet_name="농가&도축현황", startrow=21,
                     startcol=1)  # abattMonthAv 연도별 도축 개월령 현황
        df4.to_excel(writer, sheet_name="농가&도축현황", startrow=27,
                     startcol=1)  # inAbattPhenoSituation 내부 도축 형질별 추세
        df5.to_excel(writer, sheet_name="농가&도축현황", startrow=27,
                     startcol=15)  # outAbattPhenoSituation 외부 도축 형질별 추세
        # kpnDistribution KPN 사용 빈도 테이블
        df6.to_excel(writer, sheet_name="KPN사용현황", index=False)
        df7.to_excel(writer, sheet_name="KPN사용현황", startrow=30,
                     startcol=8)  # kpnEbvMeanResult 씨수소 유전능력 평균
        df8.to_excel(writer, sheet_name="유전능력현황", startrow=1,
                     startcol=1)  # breedTable 유전능력 현황(전체평균 농가평균)
        df9.to_excel(writer, sheet_name="유전능력현황", startrow=24,
                     startcol=1)  # breedTyperesult 개체별 유전능력 유형분류
        df10.to_excel(writer, sheet_name="암_도체약&육질약", index=False)
        df11.to_excel(writer, sheet_name="암_도체약&육질강", index=False)
        df12.to_excel(writer, sheet_name="암_도체강&육질약", index=False)
        df13.to_excel(writer, sheet_name="암_도체강&육질강", index=False)
        df14.to_excel(writer, sheet_name="수_도체약&육질약", index=False)
        df15.to_excel(writer, sheet_name="수_도체약&육질강", index=False)
        df16.to_excel(writer, sheet_name="수_도체강&육질약", index=False)
        df17.to_excel(writer, sheet_name="수_도체강&육질강", index=False)
        df18.to_excel(writer, sheet_name="암_도체중우선_선발지수", index=False)
        df19.to_excel(writer, sheet_name="수_도체중우선_선발지수", index=False)
        df20.to_excel(writer, sheet_name="암_근내지방우선_선발지수", index=False)
        df21.to_excel(writer, sheet_name="수_근내지방우선_선발지수", index=False)


# 칼럼명 변경
outAbattPhenoSituation.columns = [
    '외부연도별출하두수', '외부CWT평균', '외부EMA평균', '외부BFT평균', '외부MAR평균', '외부1+이상비율', '외부A', '외부B', '외부C']

fileDownload(
    farmSituationDf,
    abattSituationDf,
    abattMonthAv,
    inAbattPhenoSituation,
    outAbattPhenoSituation,
    kpnDistribution,
    kpnEbvMeanResult,
    breedTable,
    breedTyperesult,
    FgrowthImprove,
    FgrowthEnhence,
    FgQualityImprove,
    FgQualityEnhence,
    MgrowthImprove,
    MgrowthEnhence,
    MgQualityImprove,
    MgQualityEnhence,
    FselectionIndexCwt,
    MselectionIndexCwt,
    FselectionIndexMar,
    MselectionIndexMar
)

# openpyxl import

img_path = './chart'


def resizeImg(size, img_path, img_name):
    # 이미지 객체 생성
    imgs = piImage.open(img_path+"/"+img_name)
    # 이미지 resize
    resize_img = imgs.resize(size)
    # 수정한 이미지 저장
    resize_img.save(img_path+"/"+img_name, "JPEG", quality=95)


resizeImg((200, 400), img_path, '수도축개월분포_chart.jpg')
resizeImg((200, 400), img_path, '암도축개월분포_chart.jpg')
resizeImg((432, 275), img_path, 'CWT_chart.jpg')
resizeImg((432, 275), img_path, 'EMA_chart.jpg')
resizeImg((432, 275), img_path, 'BFT_chart.jpg')
resizeImg((432, 275), img_path, 'MAR_chart.jpg')
resizeImg((432, 275), img_path, '내부_chart.jpg')
resizeImg((432, 275), img_path, '외부_chart.jpg')
resizeImg((300, 600), img_path, 'CWT전체순위백분율분포_chart.jpg')
resizeImg((300, 600), img_path, 'EMA전체순위백분율분포_chart.jpg')
resizeImg((300, 600), img_path, 'BFT전체순위백분율분포_chart.jpg')
resizeImg((300, 600), img_path, 'MAR전체순위백분율분포_chart.jpg')


wb = load_workbook(f'./reportResult/{farmName[0]}_농장보고서.xlsx')


sheet1 = wb["농가&도축현황"]
sheet1.sheet_view.zoomScale = 70

sheet1SellList = [2, 3, 4, 5, 6, 7, 8, 9, 10,
                  11, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25]

for num in sheet1SellList:
    for row in range(1, 42):
        sheet1.cell(row=row, column=num).alignment = Alignment(
            horizontal='center', vertical='center')

sheet2 = wb["KPN사용현황"]
sheet2.sheet_view.zoomScale = 100

sheet2SellList_1 = [1, 2, 3, 4, 5, 6, 7]
sheet2SellList_2 = [9, 10]

for num in sheet2SellList_1:
    for row in range(1, len(sheet2['A'])+1):
        sheet2.cell(row=row, column=num).alignment = Alignment(
            horizontal='center', vertical='center')

for num in sheet2SellList_2:
    for row in range(31, 36):
        sheet2.cell(row=row, column=num).alignment = Alignment(
            horizontal='center', vertical='center')

sheet3 = wb["유전능력현황"]
sheet3.sheet_view.zoomScale = 85

sheet3SellList = [1, 2, 3, 4, 5, 6]

for num in sheet3SellList:
    for row in range(1, 31):
        sheet3.cell(row=row, column=num).alignment = Alignment(
            horizontal='center', vertical='center')


sheet4 = wb["암_도체약&육질약"]

sheet4SellList = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15]

for num in sheet4SellList:
    for row in range(1,  len(sheet4['A'])+1):
        sheet4.cell(row=row, column=num).alignment = Alignment(
            horizontal='center', vertical='center')


sheet5 = wb["암_도체약&육질강"]

sheet5SellList = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15]

for num in sheet5SellList:
    for row in range(1,  len(sheet5['A'])+1):
        sheet5.cell(row=row, column=num).alignment = Alignment(
            horizontal='center', vertical='center')

sheet6 = wb["암_도체강&육질약"]

sheet6SellList = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15]

for num in sheet6SellList:
    for row in range(1,  len(sheet6['A'])+1):
        sheet6.cell(row=row, column=num).alignment = Alignment(
            horizontal='center', vertical='center')

sheet7 = wb["암_도체강&육질강"]

sheet7SellList = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15]

for num in sheet7SellList:
    for row in range(1,  len(sheet7['A'])+1):
        sheet7.cell(row=row, column=num).alignment = Alignment(
            horizontal='center', vertical='center')

sheet8 = wb["수_도체약&육질약"]

sheet8SellList = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15]

for num in sheet8SellList:
    for row in range(1,  len(sheet8['A'])+1):
        sheet8.cell(row=row, column=num).alignment = Alignment(
            horizontal='center', vertical='center')

sheet9 = wb["수_도체약&육질강"]

sheet9SellList = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15]

for num in sheet9SellList:
    for row in range(1,  len(sheet9['A'])+1):
        sheet9.cell(row=row, column=num).alignment = Alignment(
            horizontal='center', vertical='center')

sheet10 = wb["수_도체강&육질약"]

sheet10SellList = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15]

for num in sheet10SellList:
    for row in range(1,  len(sheet10['A'])+1):
        sheet10.cell(row=row, column=num).alignment = Alignment(
            horizontal='center', vertical='center')

sheet11 = wb["수_도체강&육질강"]

sheet11SellList = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15]

for num in sheet11SellList:
    for row in range(1,  len(sheet11['A'])+1):
        sheet11.cell(row=row, column=num).alignment = Alignment(
            horizontal='center', vertical='center')

sheet12 = wb["암_도체중우선_선발지수"]

sheet12SellList = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11]

for num in sheet12SellList:
    for row in range(1,  len(sheet12['A'])+1):
        sheet12.cell(row=row, column=num).alignment = Alignment(
            horizontal='center', vertical='center')

sheet13 = wb["수_도체중우선_선발지수"]

sheet13SellList = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11]

for num in sheet13SellList:
    for row in range(1,  len(sheet13['A'])+1):
        sheet13.cell(row=row, column=num).alignment = Alignment(
            horizontal='center', vertical='center')

sheet14 = wb["암_근내지방우선_선발지수"]

sheet14SellList = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11]

for num in sheet14SellList:
    for row in range(1,  len(sheet14['A'])+1):
        sheet14.cell(row=row, column=num).alignment = Alignment(
            horizontal='center', vertical='center')

sheet15 = wb["수_근내지방우선_선발지수"]

sheet15SellList = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11]

for num in sheet15SellList:
    for row in range(1,  len(sheet15['A'])+1):
        sheet15.cell(row=row, column=num).alignment = Alignment(
            horizontal='center', vertical='center')


img1 = Image('./chart/수도축개월분포_chart.jpg')
img2 = Image('./chart/암도축개월분포_chart.jpg')
img3 = Image('./chart/CWT_chart.jpg')
img4 = Image('./chart/EMA_chart.jpg')
img5 = Image('./chart/BFT_chart.jpg')
img6 = Image('./chart/MAR_chart.jpg')
img7 = Image('./chart/내부_chart.jpg')
img8 = Image('./chart/외부_chart.jpg')
img9 = Image('./chart/KPN빈도_chart.jpg')
img10 = Image('./chart/kpnrader_chart.jpg')
img11 = Image('./chart/사육rader_chart.jpg')
img12 = Image('./chart/개체별유형분류_chart.jpg')
img13 = Image('./chart/CWT전체순위백분율분포_chart.jpg')
img14 = Image('./chart/EMA전체순위백분율분포_chart.jpg')
img15 = Image('./chart/BFT전체순위백분율분포_chart.jpg')
img16 = Image('./chart/MAR전체순위백분율분포_chart.jpg')

sheet1.add_image(img1, "H1")
sheet1.add_image(img2, "k1")
sheet1.add_image(img3, "B44")
sheet1.add_image(img4, "I44")
sheet1.add_image(img5, "B57")
sheet1.add_image(img6, "I57")
sheet1.add_image(img7, "B70")
sheet1.add_image(img8, "I70")

sheet2.add_image(img9, "I2")
sheet2.add_image(img10, "L31")

sheet3.add_image(img11, "J2")
sheet3.add_image(img12, "J24")
sheet3.add_image(img13, "B44")
sheet3.add_image(img14, "G44")
sheet3.add_image(img15, "L44")
sheet3.add_image(img16, "Q44")


print(f"{farmName[0]}_농장보고서 엑셀 생성 완료 ........")

wb.save(f'./reportResult/{farmName[0]}_농장보고서.xlsx')
